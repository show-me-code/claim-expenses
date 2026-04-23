"""
Excel Generator Module - Generate expense summary Excel file.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List

from .expense_calculator import ExpenseSummary, TripExpense


def generate_excel(summary: ExpenseSummary, output_folder: str) -> str:
    """
    Generate Excel file with expense breakdown.

    Returns path to generated Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "差旅费用汇总"

    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    money_font = Font(size=11)
    total_font = Font(bold=True, size=12)

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = '差旅费用报销汇总表'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align

    # Date
    ws['A2'] = f'生成日期: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = Font(size=10)

    # Headers
    headers = ['序号', '出差日期', '天数', '出发地', '目的地', '交通费', '住宿费', '伙食费', '合计']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill

    # Data rows
    row_num = 5
    for expense in summary.trip_expenses:
        trip = expense.trip

        ws.cell(row=row_num, column=1, value=expense.trip_id).alignment = center_align
        ws.cell(row=row_num, column=2, value=f"{trip.start_date.strftime('%Y-%m-%d')}~{trip.end_date.strftime('%Y-%m-%d') if trip.end_date else ''}").alignment = center_align
        ws.cell(row=row_num, column=3, value=trip.days).alignment = center_align
        ws.cell(row=row_num, column=4, value=trip.departure_city).alignment = left_align
        ws.cell(row=row_num, column=5, value=trip.arrival_city).alignment = left_align
        ws.cell(row=row_num, column=6, value=trip.ticket_total + trip.refund_total).alignment = right_align
        ws.cell(row=row_num, column=7, value=expense.hotel_total).alignment = right_align
        ws.cell(row=row_num, column=8, value=expense.meal_allowance).alignment = right_align
        ws.cell(row=row_num, column=9, value=expense.grand_total).alignment = right_align

        # Apply border
        for col in range(1, 10):
            ws.cell(row=row_num, column=col).border = thin_border

        row_num += 1

    # Summary row
    ws.cell(row=row_num, column=1, value='合计').font = total_font
    ws.cell(row=row_num, column=1).alignment = center_align
    ws.merge_cells(f'A{row_num}:E{row_num}')

    ws.cell(row=row_num, column=6, value=summary.total_tickets + summary.total_refunds).alignment = right_align
    ws.cell(row=row_num, column=7, value=summary.total_hotels).alignment = right_align
    ws.cell(row=row_num, column=8, value=summary.total_meals).alignment = right_align
    ws.cell(row=row_num, column=9, value=summary.grand_total).alignment = right_align

    for col in range(1, 10):
        cell = ws.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.fill = total_fill
        cell.font = total_font

    # Column widths
    widths = [6, 20, 6, 12, 12, 12, 12, 10, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Detail section
    row_num += 3
    ws.cell(row=row_num, column=1, value='详细明细').font = header_font

    for expense in summary.trip_expenses:
        trip = expense.trip
        row_num += 1

        ws.cell(row=row_num, column=1, value=f"第{expense.trip_id}次出差 ({trip.start_date.strftime('%Y-%m-%d')}~{trip.end_date.strftime('%Y-%m-%d') if trip.end_date else ''})")
        ws.cell(row=row_num, column=1).font = subheader_font

        # Ticket details
        row_num += 1
        ws.cell(row=row_num, column=2, value=f"去程: {trip.outbound_ticket.train_number if trip.outbound_ticket else ''} ￥{trip.outbound_ticket.price if trip.outbound_ticket else 0:.2f}")

        if trip.return_ticket:
            row_num += 1
            ws.cell(row=row_num, column=2, value=f"返程: {trip.return_ticket.train_number} ￥{trip.return_ticket.price:.2f}")

        for refund in trip.refund_tickets:
            row_num += 1
            ws.cell(row=row_num, column=2, value=f"退票费: {refund.train_number} ￥{refund.price:.2f}")

        # Hotel details
        for inv in expense.hotel_invoices:
            row_num += 1
            ws.cell(row=row_num, column=2, value=f"住宿: {inv.hotel_name} ({inv.days}天) ￥{inv.total:.2f}")

        row_num += 1
        ws.cell(row=row_num, column=2, value=f"伙食补助: {trip.days}天 × ￥{expense.daily_meal_rate:.2f} = ￥{expense.meal_allowance:.2f}")

        row_num += 1
        ws.cell(row=row_num, column=2, value=f"小计: ￥{expense.grand_total:.2f}")
        ws.cell(row=row_num, column=2).font = total_font

    # Save
    filename = f"差旅费用汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(output_folder, filename)
    wb.save(filepath)

    return filepath